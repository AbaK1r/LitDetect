import logging
import warnings
from pathlib import Path
from typing import List, Dict

import numpy as np
import pandas as pd
import pytorch_lightning as pl
import torch
from openpyxl import load_workbook
from tabulate import tabulate
from torch import distributed as dist
from torchmetrics.detection import MeanAveragePrecision

from .ap import ap_per_class

# 屏蔽特定警告
warnings.filterwarnings(
    "ignore",
    message="It is recommended to use `self.log"
)
logger = logging.getLogger(__name__)


class DetectionMetricsCallback(pl.Callback):
    """
    目标检测指标计算回调，支持单卡和DDP模式
    功能：
    1. 计算mAP系列指标（mAP50, mAP50-95）
    2. 计算每个类别的精确率(P)和召回率(R)
    3. 统计每个类别的图像数和实例数
    4. 输出YOLO风格的评估表格
    5. 支持DDP模式下的指标同步

    参数:
    class_names: 类别名称列表（索引必须与标签索引匹配）
    iou_threshold: 用于计算精确率和召回率的IoU阈值（默认0.5）
    """

    def __init__(self, class_names: List[str], iou_threshold: float = 0.5, save_xlsx_dir: Path = None, save_sheet_name=None):
        super().__init__()
        self.class_names = class_names
        self.num_classes = len(class_names)
        self.iou_threshold = iou_threshold
        self.save_xlsx_dir = save_xlsx_dir
        self.save_sheet_name = save_sheet_name
        self.iou_idx = None  # 存储指定IoU阈值在指标中的索引
        self.iou_50_idx = None
        self.iouv = torch.arange(0.5, 1.0, 0.05)  # IoU vector for mAP@0.5:0.95
        self.niou = self.iouv.numel()
        self.stats = None

        # 指标对象
        self.map_metric = MeanAveragePrecision(
            box_format="xyxy",  # 边界框格式：左上右下坐标
            iou_thresholds=self.iouv.tolist(),
            iou_type="bbox",  # 计算边界框IoU
            class_metrics=True,  # 计算每个类别的指标
            extended_summary=True,  # 启用详细数据（包含精确率/召回率）
            backend="pycocotools"  # 使用pycocotools后端
        )

        # 统计信息收集器
        self.class_stats = None
        self.total_images = 0

    def on_validation_epoch_start(self, trainer: pl.Trainer, pl_module: pl.LightningModule):
        """
        在每个验证epoch开始时初始化指标和统计信息
        """
        # 初始化类别统计信息张量（支持DDP同步）
        self.stats = dict(tp=[], conf=[], pred_cls=[], target_cls=[])

        # 查找指定IoU阈值在指标中的索引
        self.iou_50_idx = self._find_iou_threshold_index(pl_module, 0.5)

    def _find_iou_threshold_index(self, pl_module: pl.LightningModule, _iou_thresholds: float = 0.5) -> int:
        """
        在指标使用的IoU阈值列表中查找最接近指定阈值的索引

        返回:
        int: 最接近指定IoU阈值的索引
        """
        # 获取指标使用的IoU阈值列表
        iou_thresholds = self.map_metric.iou_thresholds

        # 转换为张量并移动到正确设备
        if not isinstance(iou_thresholds, torch.Tensor):
            iou_thresholds = torch.tensor(iou_thresholds, device=pl_module.device)

        # 计算每个阈值与指定阈值的绝对差
        diff = torch.abs(iou_thresholds - _iou_thresholds)

        # 返回最小差值的索引
        return torch.argmin(diff).item()

    def on_validation_batch_end(
            self,
            trainer: pl.Trainer,
            pl_module: pl.LightningModule,
            outputs: Dict,
            batch: Dict,
            batch_idx: int,
            dataloader_idx: int = 0
    ):
        """
        在每个验证批次结束时更新指标和收集统计信息
        """
        DEVICE = pl_module.device
        # 获取模型预测和真实标签
        preds = outputs["preds"]
        targets = outputs["targets"]

        # 更新mAP指标
        self.map_metric.update(preds, targets)

        # 更新总图像计数
        self.total_images += len(targets)

        # 处理批次中的每张图像
        for pred, target in zip(preds, targets):
            # pred:   {'boxes': tensor([m, 4]),
            #          'labels': tensor([m, ]),
            #          'scores': tensor([m, ])}
            # target: {'boxes': tensor([n, 4]),
            #          'labels': tensor([n, ])}
            n_pred = pred['boxes'].shape[0]
            n_label = target['boxes'].shape[0]
            labels = target['labels']
            image_id = target['image_id'].expand_as(labels)
            stat = dict(
                conf=torch.zeros(0, device=DEVICE),
                pred_cls=torch.zeros(0, device=DEVICE),
                tp=torch.zeros(n_pred, self.niou, dtype=torch.bool, device=DEVICE),
                target_cls=torch.stack((labels, image_id), dim=1),
            )
            if n_pred == 0:
                if n_label:
                    for k in self.stats.keys():
                        self.stats[k].append(stat[k])
                continue
            stat["conf"] = pred['scores']
            stat["pred_cls"] = pred['labels']

            if n_label:
                iou = box_iou(target['boxes'], pred['boxes'])
                stat["tp"] = self._match_predictions(pred['labels'], target['labels'], iou)
            for k in self.stats.keys():
                self.stats[k].append(stat[k])

    @staticmethod
    def _sync_tensor(_tensor, _device, _rank, _world_size):
        if _tensor.device != _device:
            _tensor = _tensor.to(_device)
        _tensor = _tensor.contiguous()

        shape_tensor = torch.tensor(_tensor.shape, device=_device, dtype=torch.long)

        shape_list = [torch.empty_like(shape_tensor) for _ in range(_world_size)]
        dist.all_gather(shape_list, shape_tensor)

        shapes = [tuple(s.tolist()) for s in shape_list]
        max_dim0 = max(shape[0] for shape in shapes)

        padded_tensor = torch.zeros((max_dim0, *_tensor.shape[1:]), device=_device, dtype=_tensor.dtype)
        padded_tensor[:_tensor.size(0)] = _tensor

        gathered_tensors = [torch.empty_like(padded_tensor) for _ in range(_world_size)]
        dist.gather(padded_tensor, gather_list=gathered_tensors if _rank == 0 else None, dst=0)

        if _rank == 0:
            # 截断填充部分
            gathered_tensors = [t[:shapes[i][0]] for i, t in enumerate(gathered_tensors)]
            gathered_tensor = torch.cat(gathered_tensors, dim=0)
            return gathered_tensor
        else:
            return None

    def get_stats(self, _device, _rank):
        # 预处理本地数据
        if len(self.stats['tp']) == 0:
            self.stats['tp'].append(torch.zeros(0, self.num_classes, dtype=torch.bool, device=_device))
        if len(self.stats['conf']) == 0:
            self.stats['conf'].append(torch.zeros(0, dtype=torch.bool, device=_device))
        if len(self.stats['pred_cls']) == 0:
            self.stats['pred_cls'].append(torch.zeros(0, dtype=torch.bool, device=_device))

        local_stats = {k: torch.cat(v, 0) for k, v in self.stats.items()}
        # print(f"\nRank {_rank}: \n", local_stats['conf'][:9], '\n', local_stats['conf'][-9:])
        if not (dist.is_available() and dist.is_initialized()):
            return local_stats

        world_size = dist.get_world_size()
        if world_size == 1:
            return local_stats

        # 同步所有统计量
        synced_stats = {}
        for k, v in local_stats.items():
            synced_tensor = self._sync_tensor(v, _device, _rank, world_size)
            if _rank == 0:
                synced_stats[k] = synced_tensor
                # print(f"Rank {_rank}: tensor {k} shape: {synced_tensor.shape}")

        # 添加同步点确保所有进程完成
        dist.barrier()

        return synced_stats if _rank == 0 else None

    def on_validation_epoch_end(self, trainer: pl.Trainer, pl_module: pl.LightningModule):
        self._record_stats(trainer, pl_module, save_results=False)

    def _record_stats(self, trainer: pl.Trainer, pl_module: pl.LightningModule, save_results=False):
        """
        在每个验证epoch结束时计算并记录指标
        """
        DEVICE = pl_module.device

        stats = self.get_stats(DEVICE, trainer.global_rank)

        # 计算mAP指标（包含扩展摘要）
        map_res = self.map_metric.compute()
        # import pickle
        # with open(f'torch.pkl', 'wb') as f:
        #     data = (self.map_metric.detection_labels,
        #             self.map_metric.detection_box,
        #             self.map_metric.detection_scores,
        #             self.map_metric.groundtruth_box)
        #     pickle.dump(data, f)
        for k, v in map_res.items():
            if isinstance(v, torch.Tensor) and len(v.shape) == 0:
                map_res[k] = v[None]
        # 从结果中提取精确率和召回率
        map_res['map_50_per_class'] = map_res['precision'][self.iou_50_idx, :, :, 0, 2].mean(dim=0)
        # map_50_95_per_class = map_res['precision'][:,:,:, 0, -1].mean(axis=(0,1))
        pl_module.log('map_50', map_res['map_50'].item(), prog_bar=True, sync_dist=True)
        # print(f"map_50 {map_res['map_50_per_class']}")
        # print(f"map_50_95 {map_50_95_per_class}")
        save_pic_dir = Path(trainer.log_dir) / 'metric_pics' / f'epoch_{trainer.current_epoch}'
        if self.save_sheet_name is not None:
            save_pic_dir = self.save_xlsx_dir.parent / self.save_sheet_name
        current_epoch = trainer.current_epoch
        if trainer.global_rank == 0:
            labels = stats['target_cls']
            images_counter = [len(torch.unique(labels[labels[:, 0] == label][:, 1])) for label in range(self.num_classes)]
            instances_counter = [labels[labels[:, 0] == label].shape[0] for label in range(self.num_classes)]
            all_images = len(torch.unique(labels[:, 1]))
            all_instances = sum(instances_counter)
            assert all_instances == labels.shape[0]
            stats['target_cls'] = stats['target_cls'][:, 0]
            stats = {k: v.cpu().numpy() for k, v in stats.items()}
            (tp, fp, p, r, f1, ap,
             unique_classes, p_curve,
             r_curve, f1_curve, x,
             prec_values) = ap_per_class(
                **stats, plot=True, names={k: v for k, v in enumerate(self.class_names)}, save_dir=save_pic_dir)
            # print(f"map_50 self {prec_values.mean(-1)}")
            tp_total = tp.sum()
            fp_total = fp.sum()
            fn_total = (stats['target_cls'] >= 0).sum() - tp_total

            _precision = tp_total / (tp_total + fp_total + 1e-8)
            _recall = tp_total / (tp_total + fn_total + 1e-8)

            # 准备表格头
            headers = ["类别", "图像数", "实例数", "TP", "FP", "精确率(P)", "召回率(R)", "AP50", "AP50-95"]
            class_data = []

            # 记录整体指标
            overall_metrics = {}
            for k, v in map_res.items():
                # 只记录关键整体指标
                if k in ['map', 'map_50', 'map_75', 'mar_100']:
                    if v.numel() == 1 and v.item() != -1:
                        overall_metrics[f"metrics/{k}"] = v.item()
            # 添加整体行到表格
            class_data.append([
                "ALL",
                all_images,
                all_instances,
                int(tp.sum()), int(fp.sum()),
                float(f"{_precision:.3f}"),  # 平均精确率
                float(f"{_recall:.3f}"),  # 平均召回率
                float(f"{map_res['map_50'].item():.3f}"),  # mAP50
                float(f"{map_res['map'].item():.3f}")  # mAP50-95
            ])

            # 准备每个类别的指标
            class_metrics = {}

            # 处理每个类别的指标
            for class_idx in range(self.num_classes):
                class_name = self.class_names[class_idx]

                # 获取统计信息
                img_count = images_counter[class_idx]
                instance_count = instances_counter[class_idx]

                # 获取精确率和召回率
                _p = p[class_idx]
                _r = r[class_idx]
                _tp = tp[class_idx]
                _fp = fp[class_idx]

                # 获取mAP指标
                map50 = map_res['map_50_per_class'][class_idx].item() if (
                        'map_50_per_class' in map_res and
                        class_idx < map_res['map_50_per_class'].numel()
                ) else 0.0

                map_all = map_res['map_per_class'][class_idx].item() if (
                        'map_per_class' in map_res and
                        class_idx < map_res['map_per_class'].numel()
                ) else 0.0

                # 添加到表格行
                class_data.append([
                    class_name,
                    img_count,
                    instance_count,
                    _tp, _fp,
                    float(f"{_p:.3f}"),
                    float(f"{_r:.3f}"),
                    float(f"{map50:.3f}"),
                    float(f"{map_all:.3f}")
                ])

                # 记录每个类别的详细指标
                class_metrics.update({
                    f"metrics-per-class/P_{class_name}": _p,
                    f"metrics-per-class/R_{class_name}": _r,
                    f"metrics-per-class/mAP50_{class_name}": map50,
                    f"metrics-per-class/mAP50-95_{class_name}": map_all,
                    # f"metrics-per-class/instances_{class_name}": instance_count,
                    # f"metrics-per-class/images_{class_name}": img_count
                })

            # 打印YOLO风格的表格
            table_str = tabulate(class_data, headers=headers, tablefmt="grid")
            logger.info(f"\n验证指标 Epoch: {current_epoch}:\n{table_str}")
            if save_results:
                self._save_metrics(class_data, headers)

            # 合并所有指标
            all_metrics = {**overall_metrics, **class_metrics}

            # 记录指标（不进行分布式同步，因为只在主进程记录）
            pl_module.log_dict(all_metrics, rank_zero_only=True)

        # 重置指标
        self._reset_metrics()

    def _save_metrics(self, class_data, headers):
        if not self.save_xlsx_dir.exists():
            sheet_num = 1
        else:
            book = load_workbook(self.save_xlsx_dir)
            sheet_num = len(book.sheetnames) + 1
        df = pd.DataFrame(class_data, columns=headers)
        new_sheet_name = self.save_sheet_name or f"model_{sheet_num}"

        if not self.save_xlsx_dir.exists():
            with pd.ExcelWriter(self.save_xlsx_dir, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=new_sheet_name, index=False)
        else:
            with pd.ExcelWriter(self.save_xlsx_dir, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        logger.info(f"指标已保存到 {new_sheet_name}")

    def _reset_metrics(self):
        """重置指标和统计信息"""
        self.map_metric.reset()
        self.stats = None

    def _match_predictions(self, pred_classes, true_classes, iou):
        """
        Matches predictions to ground truth objects (pred_classes, true_classes) using IoU.

        Args:
            pred_classes (torch.Tensor): Predicted class indices of shape(N,).
            true_classes (torch.Tensor): Target class indices of shape(M,).
            iou (torch.Tensor): An NxM tensor containing the pairwise IoU values for predictions and ground of truth

        Returns:
            (torch.Tensor): Correct tensor of shape(N,10) for 10 IoU thresholds.
        """
        # Dx10 matrix, where D - detections, 10 - IoU thresholds
        correct = np.zeros((pred_classes.shape[0], self.iouv.shape[0])).astype(bool)
        # LxD matrix where L - labels (rows), D - detections (columns)
        correct_class = true_classes[:, None] == pred_classes
        iou = iou * correct_class  # zero out the wrong classes
        iou = iou.cpu().numpy()
        for i, threshold in enumerate(self.iouv.cpu().tolist()):
            matches = np.nonzero(iou >= threshold)  # IoU > threshold and classes match
            matches = np.array(matches).T
            if matches.shape[0]:
                if matches.shape[0] > 1:
                    matches = matches[iou[matches[:, 0], matches[:, 1]].argsort()[::-1]]
                    matches = matches[np.unique(matches[:, 1], return_index=True)[1]]
                    # matches = matches[matches[:, 2].argsort()[::-1]]
                    matches = matches[np.unique(matches[:, 0], return_index=True)[1]]
                correct[matches[:, 1].astype(int), i] = True
        return torch.tensor(correct, dtype=torch.bool, device=pred_classes.device)

    def on_test_epoch_start(self, trainer: pl.Trainer, pl_module: pl.LightningModule):
        assert self.save_xlsx_dir is not None, "save_xlsx_dir must be set"
        assert self.save_sheet_name is not None, "save_sheet_name must be set"
        self.on_validation_epoch_start(trainer, pl_module)

    def on_test_batch_end(self, *args, **kwargs):
        self.on_validation_batch_end(*args, **kwargs)

    def on_test_epoch_end(self, trainer: pl.Trainer, pl_module: pl.LightningModule):
        self._record_stats(trainer, pl_module, save_results=True)


def box_iou(box1, box2, eps=1e-7):
    """
    Calculate intersection-over-union (IoU) of boxes. Both sets of boxes are expected to be in (x1, y1, x2, y2) format.
    Based on https://github.com/pytorch/vision/blob/master/torchvision/ops/boxes.py.

    Args:
        box1 (torch.Tensor): A tensor of shape (N, 4) representing N bounding boxes.
        box2 (torch.Tensor): A tensor of shape (M, 4) representing M bounding boxes.
        eps (float, optional): A small value to avoid division by zero. Defaults to 1e-7.

    Returns:
        (torch.Tensor): An NxM tensor containing the pairwise IoU values for every element in box1 and box2.
    """
    # NOTE: Need .float() to get accurate iou values
    # inter(N,M) = (rb(N,M,2) - lt(N,M,2)).clamp(0).prod(2)
    (a1, a2), (b1, b2) = box1.float().unsqueeze(1).chunk(2, 2), box2.float().unsqueeze(0).chunk(2, 2)
    inter = (torch.min(a2, b2) - torch.max(a1, b1)).clamp_(0).prod(2)

    # IoU = inter / (area1 + area2 - inter)
    return inter / ((a2 - a1).prod(2) + (b2 - b1).prod(2) - inter + eps)
